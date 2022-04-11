import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "Notice"

# headers는 본인 컴퓨터에 맞는 user-agent를 가져와서 넣어야함
headers = {"User-Agent" : "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/97.0.4692.99 Safari/537.36"}

## range()안에 수를 조정하여 페이지 수를 가져올 수 있음
for i in range(1536):
    print("시작, 페이지:",i+1)
    url = "https://ajou.ac.kr/kr/ajou/notice.do?mode=list&&articleLimit=10&article.offset={}".format(i*10)
    res = requests.get(url,headers=headers)

    # print("응답코드 : ",res.status_code)
    # if res.status_code == requests.codes.ok:
    #     print("정상입니다.")
    # else :
    #     print("문제가 생겼습니다. [에러코드 ", res.status_code, "]")

    res.raise_for_status() # raise_for_status는 문제가 생기면 바로 에러를 출력
    # print("웹 스크래핑을 진행합니다.")


    soup = BeautifulSoup(res.text,"lxml") # res.text를 lxml통해서 beautiful soup 객체로 만듦
    notices1 = soup.find_all("tr",attrs={"class":"b-top-box"})

    if i == 0:
        for k in range(len(notices1)):
            notices1_1 = notices1[k].find("a")
            notices1_2 = notices1[k].find("td",attrs={"class":"b-num-box num-notice"})
            notices1_3 = notices1_2.find("span")

            notices1_4 = (notices1_2.find_next_sibling("td"))
            notices1_5 = notices1[k].find("td",attrs={"class":"b-no-right"})
            notices1_6 = notices1_5.find_next_sibling("td")
            notices1_7 = notices1_6.find_next_sibling("td")
            ws["G{}".format(k+2)]="https://ajou.ac.kr/kr/ajou/notice.do"+notices1_1["href"]
            ws["F{}".format(k+2)]=notices1_7.get_text()
            ws["E{}".format(k+2)]=notices1_6.get_text()
            ws["D{}".format(k+2)]=notices1_5.get_text()

            ws["C{}".format(k+2)]=notices1_1["title"]
            ws["B{}".format(k+2)]=notices1_4.get_text()
            ws["A{}".format(k+2)]=notices1_3.get_text()

        ## top 공지사항 엑셀 설정
    
    notices2 = soup.find_all("tr",attrs={"class":""})
    # print(len(notices2))

    if i == 0:
        notice2_1 = notices2[0].find_all("th")
        for k in range(len(notice2_1)):
            ws.cell(row=1,column=k+1).value = notice2_1[k].get_text()
    ws["G1"] = "링크"
    ##### 여기까지가 첫 엑셀 기초 만들기
    


    for k in range(len(notices2)):
        if k == 0:
            continue
        notice = notices2[k].find("a")
        notices3 = notices2[k].find("td",attrs={'class':"b-num-box"})
        notices3_1 = notices3.find_next_sibling("td")
        notices3_2 = notices2[k].find("td",attrs={"class":"b-no-right"})
        notices3_3 = notices3_2.find_next_sibling("td")
        notices3_4 = notices3_3.find_next_sibling("td")
        
        ws["G{}".format(((len(notices2)-1)*i)+k+1+len(notices1))] = "https://ajou.ac.kr/kr/ajou/notice.do"+notice["href"]
        ws["F{}".format(((len(notices2)-1)*i)+k+1+len(notices1))] = notices3_4.get_text()
        ws["E{}".format(((len(notices2)-1)*i)+k+1+len(notices1))] = notices3_3.get_text()

        ws["D{}".format(((len(notices2)-1)*i)+k+1+len(notices1))] = notices3_2.get_text() 

        ws["C{}".format(((len(notices2)-1)*i)+k+1+len(notices1))] = notice["title"]
        ws["B{}".format(((len(notices2)-1)*i)+k+1+len(notices1))] = notices3_1.get_text()
        ws["A{}".format(((len(notices2)-1)*i)+k+1+len(notices1))]=(notices3.get_text())
        
    


    

    ## 여기까지가 notice관련된 엑셀파일 저장
    


    

wb.save("notice_all.xlsx")

